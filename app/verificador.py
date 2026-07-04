"""Ciclo gerar → criticar → revisar (padrão Reflexion).

Um REVISOR com contexto limpo (não vê o histórico do executor — evita viés de
confirmação) compara os entregáveis do workspace com a tarefa pedida. Se
reprovar, o parecer vira feedback pra uma rodada de correção.
"""

import time
from pathlib import Path

from app.config import config
from app.llm import LLM
from app.logger import logger
from app.schema import Message

_IGNORAR = {"cache", "memoria", "fila", "todo.md", "rascunhos"}
_BINARIOS = {".xlsx", ".docx", ".pptx", ".png", ".jpg", ".jpeg", ".wav", ".mp3", ".mov", ".zip"}

PROMPT_REVISOR = """Você é um revisor rigoroso de qualidade. Avalie se a tarefa foi cumprida.

TAREFA PEDIDA:
{task}

ENTREGÁVEIS PRODUZIDOS NO WORKSPACE (arquivos novos/modificados):
{arquivos}

Critérios: os arquivos pedidos existem? O conteúdo atende ao que foi pedido?
Há erros evidentes, conteúdo vazio ou placeholder?

Responda EXATAMENTE em um destes formatos:
APROVADO
ou
REPROVADO: <lista curta e objetiva dos problemas e do que corrigir>"""


def _entregaveis(desde: float) -> list[Path]:
    out = []
    for p in config.workspace_root.rglob("*"):
        if not p.is_file():
            continue
        rel = p.relative_to(config.workspace_root)
        if any(part in _IGNORAR for part in rel.parts):
            continue
        if p.stat().st_mtime >= desde:
            out.append(p)
    return sorted(out)


def _preview(p: Path, limite: int = 1200) -> str:
    if p.suffix.lower() in _BINARIOS:
        extra = ""
        try:
            if p.suffix.lower() == ".xlsx":
                from openpyxl import load_workbook

                ws = load_workbook(p, read_only=True).active
                linhas = [[c.value for c in row] for row in ws.iter_rows(max_row=5)]
                extra = f" | primeiras linhas: {linhas}"
            elif p.suffix.lower() == ".docx":
                from docx import Document

                paras = [q.text for q in Document(str(p)).paragraphs[:5] if q.text]
                extra = f" | início: {paras}"
        except Exception:
            pass
        return f"(arquivo binário, {p.stat().st_size} bytes{extra})"
    try:
        if p.suffix.lower() == ".pdf":
            import pdfplumber

            with pdfplumber.open(p) as pdf:
                return (pdf.pages[0].extract_text() or "")[:limite]
        return p.read_text(encoding="utf-8", errors="replace")[:limite]
    except Exception as e:
        return f"(não foi possível ler: {e})"


async def revisar(task: str, desde: float) -> tuple[bool, str]:
    """Avalia os entregáveis com um LLM de contexto limpo. (aprovado, parecer)."""
    arquivos = _entregaveis(desde)
    listagem = (
        "\n\n".join(f"--- {p.relative_to(config.workspace_root)} ---\n{_preview(p)}" for p in arquivos[:10])
        or "(NENHUM arquivo novo ou modificado no workspace)"
    )
    llm = LLM()
    parecer = (
        await llm.ask(
            [Message.user_message(PROMPT_REVISOR.format(task=task, arquivos=listagem))],
            stream=False,
        )
    ).strip()
    aprovado = parecer.upper().lstrip("* #").startswith("APROVADO")
    return aprovado, parecer


async def executar_com_verificacao(
    task: str, max_steps: int = 10, rodadas_correcao: int = 1
) -> tuple[bool, str]:
    """Executa a tarefa, revisa e corrige até aprovar (ou esgotar as rodadas)."""
    from app.agent.mangaba import Mangaba

    inicio = time.time()
    agent = await Mangaba.create(max_steps=max_steps)
    try:
        await agent.run(task)
    finally:
        await agent.cleanup()

    aprovado, parecer = await revisar(task, inicio)
    rodada = 0
    while not aprovado and rodada < rodadas_correcao:
        rodada += 1
        logger.warning(f"🔍 Revisor reprovou (rodada {rodada}): {parecer[:300]}")
        corretor = await Mangaba.create(max_steps=max_steps)
        try:
            await corretor.run(
                "Uma tarefa foi executada mas o REVISOR apontou problemas. "
                f"Corrija os entregáveis no workspace e finalize.\n\nTAREFA ORIGINAL: {0}\n\nPARECER DO REVISOR: {1}".format(
                    task, parecer
                )
            )
        finally:
            await corretor.cleanup()
        aprovado, parecer = await revisar(task, inicio)

    if aprovado:
        logger.info("✅ Revisor aprovou os entregáveis")
    else:
        logger.error(f"❌ Reprovado após {rodadas_correcao} correção(ões): {parecer[:300]}")
    return aprovado, parecer
